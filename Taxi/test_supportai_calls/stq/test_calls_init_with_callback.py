import datetime
import io
import json
from typing import List

import openpyxl
import pandas as pd
import pytest

from supportai_lib.tasks import base_task
from supportai_lib.tasks import constants

from supportai_calls import models as db_models
from supportai_calls.utils import call_helpers
from test_supportai_calls import common

pytestmark = [  # pylint: disable=invalid-name
    pytest.mark.config(
        TVM_RULES=[
            {'src': 'supportai-calls', 'dst': 'personal'},
            {'src': 'supportai-calls', 'dst': 'stq-agent'},
            {'src': 'supportai-calls', 'dst': 'supportai-context'},
            {'src': 'supportai-calls', 'dst': 'supportai-tasks'},
        ],
    ),
]

SHOULD_BE_ENDED_ERROR_CODE = call_helpers.constants.ABONENT_HANGUP_CODE
SHOULD_BE_DELAYED_ERROR_CODE = call_helpers.constants.ABONENT_REFUSED_CODE
BOTH_STATUSES_ERROR_CODE = 'some error'


def _get_file_with_calls(calls: List[db_models.Call]) -> bytes:
    workbook = openpyxl.Workbook()
    phones_sheet = workbook.active
    phones_sheet.title = 'phones'
    phones_sheet.cell(1, 1, 'phone')
    for idx, call in enumerate(calls):
        phones_sheet.cell(idx + 2, 1, call.phone)

    data_stream = io.BytesIO()
    workbook.save(data_stream)
    workbook.close()
    data_stream.seek(0)

    return data_stream.read()


@pytest.mark.parametrize('calls_from_db', [False, True])
@pytest.mark.parametrize(
    ('callback_after', 'num_attempts'), [(None, None), (3, None), (3, 5)],
)
@pytest.mark.project(slug='sample_project')
async def test_decrementing_num_attempts_when_initiating(
        stq3_context,
        stq_runner,
        mockserver,
        calls_from_db,
        callback_after,
        num_attempts,
        create_task,
        create_task_file,
):
    now = datetime.datetime.now()
    real_num_attempts = num_attempts or 1

    task_params = {}
    if calls_from_db:
        task_params['calls_from_db'] = True
    if callback_after is not None:
        task_params['callback_after'] = callback_after
    if num_attempts is not None:
        task_params['num_attempts'] = num_attempts

    task = create_task(type_='outgoing_calls_init', params=task_params)

    calls = [
        common.get_preset_call(
            'sample_project',
            phone=f'+{idx}',
            created=now,
            task_id=task.id,
            attempts_left=real_num_attempts,
        )
        for idx in range(2)
    ]

    async with stq3_context.pg.master_pool.acquire() as conn:
        if calls_from_db:
            await db_models.Call.insert_bulk(stq3_context, conn, calls)
        else:
            file_data = _get_file_with_calls(calls)
            file = create_task_file(
                filename='he-he.xlsx',
                content_type=constants.XLSX_CONTENT_TYPE,
                content=file_data,
            )
            task.file_id = file.id

    @mockserver.json_handler('/personal/v1/phones/bulk_store')
    def _(_):
        return {
            'items': [
                {'id': 'some_id', 'value': f'+{idx}'} for idx in range(2)
            ],
        }

    @mockserver.json_handler('/ivr-dispatcher/v1/ivr-framework/create-call')
    async def _(_):
        return {}

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id, args=(), kwargs={'task_id': task.id},
    )
    assert task.status == base_task.TaskStatus.PROCESSING.value

    async with stq3_context.pg.slave_pool.acquire() as conn:
        calls = await db_models.Call.select_by_task_id(
            stq3_context, conn, task.id,
        )

    assert len(calls) == 2

    for call in calls:
        if call.attempts_left is not None:
            assert call.attempts_left == real_num_attempts - 1


@pytest.mark.project(slug='sample_project')
@pytest.mark.parametrize('num_iterations', [1, 2])
async def test_delay_when_update_calls(
        stq3_context, stq_runner, mockserver, create_task, num_iterations,
):
    now = datetime.datetime.now()
    callback_after_in_minutes = 2

    attempts_left_cases = [1, 0, None]
    errors = [
        None,
        SHOULD_BE_ENDED_ERROR_CODE,
        SHOULD_BE_DELAYED_ERROR_CODE,
        BOTH_STATUSES_ERROR_CODE,
    ]
    error_to_expected_status = {
        SHOULD_BE_ENDED_ERROR_CODE: db_models.CallStatus.ENDED,
        SHOULD_BE_DELAYED_ERROR_CODE: db_models.CallStatus.DELAYED,
        BOTH_STATUSES_ERROR_CODE: None,
    }

    calls = []
    personal_id_to_expected_status = {}
    personal_id_to_chat_id = {}
    chat_id_to_error = {}
    idx = 0

    task = create_task(
        type_='outgoing_calls_init',
        status=base_task.TaskStatus.PROCESSING,
        state=base_task.TaskProcessingState.Working,
    )
    for attempts_left in attempts_left_cases:
        for error in errors:
            idx += 1
            personal_phone_id = f'id{idx}'
            chat_id = f'chat_id{idx}'

            calls.append(
                db_models.Call(
                    id=-1,
                    project_slug='sample_project',
                    direction=db_models.CallDirection.OUTGOING,
                    call_service=db_models.CallService.IVR_FRAMEWORK,
                    task_id=task.id,
                    chat_id=chat_id,
                    phone=str(idx),
                    personal_phone_id=personal_phone_id,
                    features='{}',
                    attempt_number=1,
                    status=db_models.CallStatus.INITIATED,
                    has_record=False,
                    created=now,
                    initiated=now,
                    attempts_left=attempts_left,
                ),
            )
            expected_status = db_models.CallStatus.PROCESSING
            if error is not None:
                expected_status = error_to_expected_status[error]
            if expected_status == db_models.CallStatus.DELAYED:
                if attempts_left is None or attempts_left == 0:
                    expected_status = db_models.CallStatus.ENDED
            if expected_status is None:
                if (
                        attempts_left is None
                        or attempts_left == 0
                        or num_iterations == 2
                ):
                    expected_status = db_models.CallStatus.ERROR
                else:
                    expected_status = db_models.CallStatus.DELAYED

            personal_id_to_expected_status[personal_phone_id] = expected_status
            personal_id_to_chat_id[personal_phone_id] = chat_id
            chat_id_to_error[chat_id] = error

    async with stq3_context.pg.master_pool.acquire() as conn:
        await db_models.Call.insert_bulk(stq3_context, conn, calls)

    @mockserver.json_handler('/supportai-context/v1/contexts/multiple')
    async def _(_):
        contexts = []
        for chat_id_, error_ in chat_id_to_error.items():
            contexts.append(
                common.create_context(
                    chat_id_,
                    error_code=error_,
                    error_code_index=num_iterations - 1,
                    num_records=2,
                    created_at=now,
                    is_ended=False,
                ),
            )
        return {'contexts': contexts, 'total': len(chat_id_to_error)}

    await stq_runner.supportai_calls_processing.call(
        task_id=2,
        args=(),
        kwargs={
            'task_args': {'callback_after': callback_after_in_minutes},
            'task_id': task.id,
        },
    )

    async with stq3_context.pg.slave_pool.acquire() as conn:
        updated_calls = await db_models.Call.select_by_task_id(
            stq3_context, conn, task_id=task.id,
        )

    assert len(updated_calls) == len(calls)

    assert task.status != base_task.TaskStatus.ERROR.value, task.error_message

    for updated_call in updated_calls:
        personal_phone_id = updated_call.personal_phone_id
        assert (
            updated_call.status
            == personal_id_to_expected_status[personal_phone_id]
        )
        if updated_call.status == db_models.CallStatus.DELAYED:
            assert (
                updated_call.callback_at - updated_call.initiated
            ).seconds >= callback_after_in_minutes * 60
            chat_id = updated_call.chat_id
            assert chat_id == personal_id_to_chat_id[personal_phone_id]


@pytest.mark.config(
    TVM_RULES=[
        {'src': 'supportai-calls', 'dst': 'personal'},
        {'src': 'supportai-calls', 'dst': 'stq-agent'},
        {'src': 'supportai-calls', 'dst': 'supportai-context'},
        {'src': 'supportai-calls', 'dst': 'supportai-tasks'},
    ],
)
async def test_order_of_delayed_when_getting_calls(
        stq3_context, stq_runner, mockserver, create_task,
):

    now = datetime.datetime.now()
    delays_in_minutes = [1, 7, 1, 5, 6, -3, -1, 1, -2]
    id_delay_pairs = [
        (idx, delay)
        for idx, delay in enumerate(delays_in_minutes)
        if delay < 0
    ]
    phones_in_order = [
        f'{item[0]}' for item in sorted(id_delay_pairs, key=lambda x: x[1])
    ]

    task = create_task(
        type_='outgoing_calls_init',
        status=base_task.TaskStatus.PROCESSING,
        state=base_task.TaskProcessingState.Working,
    )

    calls = [
        db_models.Call(
            id=-1,
            project_slug='sample_project',
            direction=db_models.CallDirection.OUTGOING,
            call_service=db_models.CallService.IVR_FRAMEWORK,
            task_id=task.id,
            chat_id=f'chat_id{i}',
            phone=f'{i}',
            personal_phone_id=f'{i}',
            features='{}',
            attempt_number=1,
            status=db_models.CallStatus.DELAYED,
            has_record=False,
            created=now,
            callback_at=now + datetime.timedelta(minutes=delay),
        )
        for i, delay in enumerate(delays_in_minutes)
    ]

    async with stq3_context.pg.master_pool.acquire() as conn:
        await db_models.Call.insert_bulk(stq3_context, conn, calls)

    @mockserver.json_handler('/ivr-dispatcher/v1/ivr-framework/create-call')
    async def mlw_handler(request):
        phone_number = request.json['actions'][0]['originate']['phone_number']
        assert phone_number == phones_in_order.pop(0)
        return {}

    await stq_runner.supportai_calls_processing.call(
        task_id=3,
        args=(),
        kwargs={'task_id': task.id, 'task_args': {'callback_after': 1}},
    )
    assert task.status == base_task.TaskStatus.PROCESSING.value

    assert mlw_handler.times_called == 3


@pytest.mark.config(
    TVM_RULES=[
        {'src': 'supportai-calls', 'dst': 'personal'},
        {'src': 'supportai-calls', 'dst': 'stq-agent'},
        {'src': 'supportai-calls', 'dst': 'supportai-context'},
        {'src': 'supportai-calls', 'dst': 'supportai-tasks'},
    ],
)
@pytest.mark.project(slug='sample_project')
async def test_mixing_calls_when_getting_calls(
        stq3_context, stq_runner, mockserver, create_task,
):
    now = datetime.datetime.now().astimezone()
    statuses = [db_models.CallStatus.DELAYED] * 3 + [
        db_models.CallStatus.QUEUED,
    ] * 3

    task = create_task(
        type_='outgoing_calls_init',
        status=base_task.TaskStatus.PROCESSING,
        state=base_task.TaskProcessingState.Working,
    )

    calls = [
        db_models.Call(
            id=-1,
            project_slug='sample_project',
            direction=db_models.CallDirection.OUTGOING,
            call_service=db_models.CallService.IVR_FRAMEWORK,
            task_id=task.id,
            attempt_number=1,
            chat_id=f'initial_chat_id_{i}',
            phone=f'phone_{i}',
            personal_phone_id=f'{i}',
            features='{}',
            status=status,
            has_record=False,
            created=now + datetime.timedelta(minutes=i),
            callback_at=now + datetime.timedelta(minutes=-i - 1)
            if status == db_models.CallStatus.DELAYED
            else None,
        )
        for i, status in enumerate(statuses)
    ]

    async with stq3_context.pg.master_pool.acquire() as conn:
        await db_models.Call.insert_bulk(stq3_context, conn, calls)

    chat_ids_in_order = [
        'phone_2',
        'phone_1',
        'initial_chat_id_3',
        'initial_chat_id_4',
    ]

    @mockserver.json_handler('/ivr-dispatcher/v1/ivr-framework/create-call')
    async def mlw_handler(request):
        chat_id = request.json['call_external_id']
        if len(chat_ids_in_order) > 2:
            assert chat_id[:7] == chat_ids_in_order.pop(0)
        else:
            assert chat_id == chat_ids_in_order.pop(0)
        return {}

    await stq_runner.supportai_calls_processing.call(
        task_id=4,
        args=(),
        kwargs={'task_id': task.id, 'task_args': {'callback_after': 3}},
    )

    assert mlw_handler.times_called == 4


@pytest.mark.parametrize('callback_after', [0, 1])
@pytest.mark.project(slug='sample_project')
async def test_order_of_refreshing_delayed_call_state(
        stq3_context, stq_runner, mockserver, create_task, callback_after,
):
    now = datetime.datetime.now().astimezone()
    attempts_left_list = [None, 0, 1]
    error_code = 'WOW ERROR'

    task = create_task(
        type_='outgoing_calls_init',
        status=base_task.TaskStatus.PROCESSING,
        state=base_task.TaskProcessingState.Working,
    )

    calls = [
        common.get_preset_call(
            project_slug='sample_project',
            chat_id=str(idx),
            task_id=task.id,
            attempts_left=attempts_left,
            attempt_number=1,
            status=db_models.CallStatus.INITIATED,
            created=now,
            initiated=now,
        )
        for idx, attempts_left in enumerate(attempts_left_list)
    ]

    async with stq3_context.pg.master_pool.acquire() as conn:
        await db_models.Call.insert_bulk(stq3_context, conn, calls)

    @mockserver.json_handler('supportai-context/v1/contexts/multiple')
    async def _(_):
        return {
            'contexts': [
                common.create_context(
                    chat_id=str(idx),
                    num_records=10,
                    error_code=error_code,
                    error_code_index=0,
                )
                for idx in range(3)
            ],
            'total': 2,
        }

    @mockserver.json_handler('ivr-dispatcher/v1/ivr-framework/create-call')
    async def _(_):
        return {}

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id,
        args=(),
        kwargs={
            'task_id': task.id,
            'task_args': {'callback_after': callback_after},
        },
    )

    async with stq3_context.pg.slave_pool.acquire() as conn:
        updated_calls = await db_models.Call.select_by_task_id(
            stq3_context, conn, task.id,
        )

    error_calls = [
        call for call in updated_calls if call.chat_id in ['0', '1']
    ]
    delayed_call = [
        call for call in updated_calls if call.chat_id not in ['0', '1']
    ][0]

    error_message = f'Ошибка инициации звонка: {error_code}'

    for error_call in error_calls:
        assert error_call.status == db_models.CallStatus.ERROR
        assert error_call.error_message == error_message
        assert error_call.ended is not None

    if callback_after == 0:
        assert delayed_call.attempts_left == 0
        assert delayed_call.status == db_models.CallStatus.INITIATED
        assert delayed_call.error_message is None
        assert delayed_call.ended is None
    else:
        assert delayed_call.attempts_left == 1
        assert delayed_call.status == db_models.CallStatus.DELAYED
        assert delayed_call.error_message == error_message
        assert delayed_call.ended is not None

    assert delayed_call.callback_at == now + datetime.timedelta(
        minutes=callback_after,
    )


@pytest.mark.project(slug='sample_project')
async def test_delaying_calls_with_no_context(
        stq3_context, stq_runner, mockserver, create_task,
):
    @mockserver.json_handler('supportai-context/v1/contexts/multiple')
    async def _(_):
        return {'contexts': [], 'total': 0}

    attempts_left_list = [None, 0, 1]
    now = datetime.datetime.now().astimezone()

    task = create_task(
        type_='outgoing_calls_init',
        status=base_task.TaskStatus.PROCESSING,
        state=base_task.TaskProcessingState.Working,
    )

    calls = [
        common.get_preset_call(
            project_slug='sample_project',
            chat_id=f'chat_id_{idx}',
            task_id=task.id,
            created=now,
            initiated=now - datetime.timedelta(seconds=120),
            attempt_number=1,
            status=db_models.CallStatus.PROCESSING,
            attempts_left=attempts_left,
        )
        for idx, attempts_left in enumerate(attempts_left_list)
    ]

    async with stq3_context.pg.master_pool.acquire() as conn:
        await db_models.Call.insert_bulk(stq3_context, conn, calls)

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id,
        args=(),
        kwargs={'task_id': task.id, 'task_args': {'callback_after': 4}},
    )

    async with stq3_context.pg.slave_pool.acquire() as conn:
        updated_calls = await db_models.Call.select_by_task_id(
            stq3_context, conn, task.id,
        )

    error_message = (
        'The call has been initiated but does not have dialog history'
    )
    for call in updated_calls:
        assert call.error_message == error_message
        if call.chat_id == 'chat_id_2':
            assert call.status == db_models.CallStatus.DELAYED
        else:
            assert call.status == db_models.CallStatus.ERROR


@pytest.mark.project(slug='sample_project')
async def test_delay_call_by_tag(
        stq3_context, stq_runner, mockserver, create_task,
):
    num_attempts = 3
    callback_after_in_minutes = 4

    calls_cases = [
        # chat_id, error, has callback tag, delay_minutes_feature
        ('1', 'dial error', False, None),
        ('2', 'dial error', True, None),
        ('3', None, False, None),
        ('4', None, True, None),
        ('5', None, True, 3),
    ]

    @mockserver.json_handler('supportai-context/v1/contexts/multiple')
    async def _(_):
        return {
            'contexts': [
                common.create_context(
                    chat_id=chat_id,
                    num_records=10,
                    error_code=error,
                    error_code_index=0 if error else None,
                    tags=['callback'] if has_callback_tag else None,
                    responses_features=[
                        [{'key': 'callback_delay_in_minutes', 'value': delay}]
                        * 10,
                    ]
                    if delay
                    else None,
                )
                for chat_id, error, has_callback_tag, delay in calls_cases
            ],
            'total': 5,
        }

    task = create_task(
        type_='outgoing_calls_init',
        status=base_task.TaskStatus.PROCESSING,
        state=base_task.TaskProcessingState.Working,
    )

    calls = [
        common.get_preset_call(
            project_slug='sample_project',
            chat_id=chat_id,
            task_id=task.id,
            attempts_left=num_attempts,
            status=db_models.CallStatus.INITIATED,
            attempt_number=1,
            initiated=datetime.datetime.now(),
        )
        for chat_id, *_ in calls_cases
    ]

    async with stq3_context.pg.master_pool.acquire() as conn:
        await db_models.Call.insert_bulk(stq3_context, conn, calls)

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id,
        args=(),
        kwargs={
            'task_id': task.id,
            'task_args': {'callback_after': callback_after_in_minutes},
        },
    )
    assert (
        task.status == base_task.TaskStatus.PROCESSING.value
    ), task.error_message

    async with stq3_context.pg.slave_pool.acquire() as conn:
        updated_calls = await db_models.Call.select_by_task_id(
            stq3_context, conn, task.id,
        )
    assert len(updated_calls) == 5

    for updated_call in updated_calls:
        if updated_call.chat_id in ['2', '4', '5']:
            assert updated_call.attempts_left == 1
        else:
            assert updated_call.attempts_left == 3

        if updated_call.chat_id == '3':
            assert updated_call.status == db_models.CallStatus.ENDED
            assert updated_call.callback_at is None
        else:
            assert updated_call.status == db_models.CallStatus.DELAYED
            assert updated_call.callback_at is not None
            delay_seconds = int(
                (
                    updated_call.callback_at - updated_call.ended
                ).total_seconds(),
            )
            expected_delay_minutes = (
                3 if updated_call.chat_id == '5' else callback_after_in_minutes
            )
            assert delay_seconds == expected_delay_minutes * 60


@pytest.mark.project(slug='sample_project')
async def test_delay_only_by_last_tagged_dialog_records(
        stq3_context, stq_runner, mockserver, create_task,
):
    chat_id_to_idx_from_the_end = {str(idx): idx for idx in range(1, 6)}

    task = create_task(
        type_='outgoing_calls_init',
        status=base_task.TaskStatus.PROCESSING,
        state=base_task.TaskProcessingState.Working,
    )

    calls = [
        common.get_preset_call(
            project_slug='sample_project',
            chat_id=str(idx),
            task_id=task.id,
            attempts_left=3,
            status=db_models.CallStatus.INITIATED,
            attempt_number=1,
            initiated=datetime.datetime.now(),
        )
        for idx in range(1, 6)
    ]
    async with stq3_context.pg.master_pool.acquire() as conn:
        await db_models.Call.insert_bulk(stq3_context, conn, calls)

    @mockserver.json_handler('supportai-context/v1/contexts/multiple')
    async def _(_):
        return {
            'contexts': [
                common.create_context(
                    chat_id=chat_id,
                    num_records=10,
                    tags=['callback'],
                    tags_index=10 - idx,
                )
                for chat_id, idx in chat_id_to_idx_from_the_end.items()
            ],
            'total': 5,
        }

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id,
        args=(),
        kwargs={'task_id': task.id, 'task_args': {'callback_after': 5}},
    )
    assert task.status == base_task.TaskStatus.PROCESSING.value

    async with stq3_context.pg.slave_pool.acquire() as conn:
        updated_calls = await db_models.Call.select_by_task_id(
            stq3_context, conn, task.id,
        )
    assert len(updated_calls) == 5

    for updated_call in updated_calls:
        idx_from_the_end = chat_id_to_idx_from_the_end[updated_call.chat_id]
        if idx_from_the_end > 3:
            assert updated_call.status == db_models.CallStatus.ENDED
        else:
            assert updated_call.status == db_models.CallStatus.DELAYED


@pytest.mark.parametrize('callback_after', [5, 10])
async def test_cancel_too_late_delayed_calls(
        stq3_context, stq_runner, mockserver, create_task, callback_after,
):
    now = datetime.datetime.now().astimezone()

    task = create_task(
        type_='outgoing_calls_init',
        status=base_task.TaskStatus.PROCESSING,
        state=base_task.TaskProcessingState.Working,
    )

    calls = [
        common.get_preset_call(
            project_slug='sample_project',
            chat_id='chat_id',
            task_id=task.id,
            status=db_models.CallStatus.INITIATED,
            attempt_number=1,
            initiated=now,
            attempts_left=5,
            end_at=now + datetime.timedelta(minutes=8),
        ),
    ]
    async with stq3_context.pg.master_pool.acquire() as conn:
        await db_models.Call.insert_bulk(stq3_context, conn, calls)

    @mockserver.json_handler('supportai-context/v1/contexts/multiple')
    async def _(_):
        return {
            'contexts': [
                common.create_context(
                    chat_id='chat_id',
                    error_code='some error',
                    error_code_index=0,
                    num_records=10,
                    created_at=now,
                ),
            ],
            'total': 1,
        }

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id,
        args=(),
        kwargs={
            'task_id': task.id,
            'task_args': {'callback_after': callback_after},
            'calls_from_db': True,
            'end_calls_at': str(now),
        },
    )
    if callback_after == 5:
        assert task.status == base_task.TaskStatus.PROCESSING.value
    else:
        assert task.status == base_task.TaskStatus.COMPLETED.value

    async with stq3_context.pg.slave_pool.acquire() as conn:
        updated_calls = await db_models.Call.select_by_task_id(
            stq3_context, conn, task.id,
        )
    assert len(updated_calls) == 1

    if callback_after == 5:
        assert updated_calls[0].status == db_models.CallStatus.DELAYED
    else:
        assert updated_calls[0].status == db_models.CallStatus.CANCELLED


@pytest.mark.parametrize('attempts', [1, 0])
@pytest.mark.project(slug='project_voximplant')
@pytest.mark.pgsql('supportai_calls', files=['sample_project.sql'])
async def test_delaying_calls_in_voximplant(
        stq3_context, stq_runner, mockserver, create_task, attempts,
):
    chat_id_to_case = {
        # error_code, call_connected, expected_status
        '0': ('ABONENT_HANGUP', None, 'ended'),
        '1': ('UNKNOWN_ERROR', True, 'error'),
        '2': ('DESTINATION_BUSY', False, 'delayed' if attempts else 'ended'),
        '3': ('UNKNOWN_ERROR', False, 'delayed' if attempts else 'error'),
        '4': (None, True, 'ended'),
    }

    now = datetime.datetime.now().astimezone()
    call_ended_at = now + datetime.timedelta(minutes=5)

    task = create_task(
        type_='outgoing_calls_init',
        status=base_task.TaskStatus.PROCESSING,
        state=base_task.TaskProcessingState.Working,
        params={'list_id': 1},
    )

    calls = [
        common.get_preset_call(
            project_slug='project_voximplant',
            call_service=db_models.CallService.VOXIMPLANT,
            chat_id=chat_id,
            task_id=task.id,
            status=db_models.CallStatus.INITIATED,
            attempt_number=1,
            initiated=now,
            attempts_left=attempts,
        )
        for chat_id in chat_id_to_case
    ]
    async with stq3_context.pg.master_pool.acquire() as conn:
        await db_models.Call.insert_bulk(stq3_context, conn, calls)

    @mockserver.json_handler('voximplant/platform_api/GetCallListDetails')
    async def _(_):
        results = []
        for (
                chat_id,
                (error_code, call_connected_, _),
        ) in chat_id_to_case.items():
            result = {
                'start_execution_time': '',
                'attempts_left': 1,
                'status_id': 2,
                'list_id': 123,
                'finish_execution_time': '',
                'start_at': '',
                'custom_data': json.dumps({'chat_id': chat_id}),
                'last_attempt': now.isoformat(),
                'status': 'Processed',
            }
            result_data = {
                'started_at': now.isoformat(),
                'ended_at': call_ended_at.isoformat(),
            }
            if call_connected_ is not None:
                result_data['call_connected'] = call_connected_
            if error_code:
                result_data['error_code'] = error_code
            result['result_data'] = json.dumps(result_data)
            results.append(result)

        return {'result': results, 'count': 5}

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id,
        args=(),
        kwargs={
            'task_id': task.id,
            'task_args': {'callback_after': 5, 'list_id': '1'},
            'calls_from_db': True,
        },
    )
    if attempts == 1:
        assert task.status == base_task.TaskStatus.PROCESSING.value
    else:
        assert task.status == base_task.TaskStatus.COMPLETED.value

    async with stq3_context.pg.slave_pool.acquire() as conn:
        updated_calls = await db_models.Call.select_by_task_id(
            stq3_context, conn, task.id,
        )
    assert len(updated_calls) == 5
    for call in updated_calls:
        error, call_connected, expected_status = chat_id_to_case[call.chat_id]
        assert call.status.value == expected_status
        if expected_status == 'delayed':
            assert call.callback_at == now + datetime.timedelta(minutes=10)
        elif expected_status == 'error':
            abstract = (
                'Ошибка во время разговора: '
                if call_connected
                else 'Ошибка инициации звонка: '
            )
            assert call.error_message == abstract + error


@pytest.mark.parametrize('list_id', [123, None])
@pytest.mark.project(slug='project_voximplant')
@pytest.mark.pgsql('supportai_calls', files=['sample_project.sql'])
async def test_getting_next_calls_voximplant(
        stq3_context, stq_runner, mockserver, create_task, list_id,
):
    task_params = {'list_id': list_id} if list_id else {}
    task = create_task(
        type_='outgoing_calls_init',
        status=base_task.TaskStatus.PROCESSING,
        state=base_task.TaskProcessingState.Working,
        params=task_params,
    )
    calls = [
        common.get_preset_call(
            project_slug='project_voximplant',
            call_service=db_models.CallService.VOXIMPLANT,
            chat_id=str(idx),
            phone=str(idx),
            task_id=task.id,
            status=status,
            attempt_number=1,
            attempts_left=1,
        )
        for idx, status in enumerate(
            (db_models.CallStatus.QUEUED, db_models.CallStatus.DELAYED),
        )
    ]
    async with stq3_context.pg.master_pool.acquire() as conn:
        await db_models.Call.insert_bulk(stq3_context, conn, calls)

    updated_chat_ids = {}

    @mockserver.json_handler('voximplant/platform_api/CreateCallList')
    async def create_handler(request):
        if list_id:
            assert False

        nonlocal updated_chat_ids

        phones_df = pd.read_csv(
            io.BytesIO(request.get_data()), sep=';', dtype=str,
        )
        assert set(phones_df['number']) == {'0', '1'}
        updated_chat_ids = set(phones_df['chat_id'])
        return {'result': True, 'list_id': 123, 'count': 2}

    @mockserver.json_handler('voximplant/platform_api/AppendToCallList')
    async def append_handler(request):
        if not list_id:
            assert False

        nonlocal updated_chat_ids

        phones_df = pd.read_csv(
            io.BytesIO(request.get_data()), sep=';', dtype=str,
        )
        updated_chat_ids = set(phones_df['chat_id'])
        assert set(phones_df['number']) == {'0', '1'}
        return {'result': True, 'list_id': 123, 'count': 2}

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id,
        args=(),
        kwargs={
            'task_id': task.id,
            'task_args': {'callback_after': 5, 'calls_from_db': True},
        },
    )

    assert (
        task.status == base_task.TaskStatus.PROCESSING.value
    ), task.error_message
    assert create_handler.times_called + append_handler.times_called == 1

    async with stq3_context.pg.slave_pool.acquire() as conn:
        updated_calls = await db_models.Call.select_by_task_id(
            stq3_context, conn, task.id,
        )
    assert len(updated_calls) == 2
    assert set(call.chat_id for call in updated_calls) == updated_chat_ids
    for call in updated_calls:
        assert call.attempts_left == 0
