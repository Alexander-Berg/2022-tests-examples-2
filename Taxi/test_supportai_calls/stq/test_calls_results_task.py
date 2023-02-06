import copy
import datetime
import io
import random
from typing import Dict
from typing import List

import aiohttp
from aiohttp import web
import openpyxl
import pandas as pd
import pytest

from supportai_lib.tasks import base_task

from supportai_calls.generated.web.error_middleware import plugin as errors
import supportai_calls.models as db_models
from supportai_calls.stq import calls_processing
from supportai_calls.stq.tasks import outgoing_calls_results_task
from supportai_calls.utils import constants
from test_supportai_calls import common

FILE_EXTENSION_TO_CONTENT_TYPE = {
    'xlsx': constants.XLSX_CONTENT_TYPE,
    'xlsm': constants.XLSM_CONTENT_TYPE,
}

pytestmark = [  # pylint: disable=invalid-name
    pytest.mark.pgsql('supportai_calls', files=['sample_project.sql']),
    pytest.mark.config(
        TVM_RULES=[
            {'src': 'supportai-calls', 'dst': 'supportai-context'},
            {'src': 'supportai-calls', 'dst': 'stq-agent'},
            {'src': 'supportai-calls', 'dst': 'supportai-tasks'},
        ],
    ),
]


def _convert_list_feature_to_string(feature_list_value: List) -> str:
    converted_values_list = []
    for elem in feature_list_value:
        converted_values_list.append(
            f'"{elem}"' if isinstance(elem, str) else str(elem),
        )
    return ', '.join(converted_values_list)


def _check_table_header_beginning(file, expected_header: List[str]):
    file_data: pd.DataFrame = pd.read_excel(
        io.BytesIO(file), sheet_name=None, na_filter=False,
    )
    parsed_df = file_data['Parsed Calls Data']
    assert list(parsed_df.columns)[: len(expected_header)] == expected_header


def _check_features_in_file(file, expected_features: Dict):
    file_data: pd.DataFrame = pd.read_excel(
        io.BytesIO(file), sheet_name=None, na_filter=False,
    )
    parsed_df = file_data['Parsed Calls Data']
    records = parsed_df.to_dict('records')
    phone_to_record = {record['Телефон']: record for record in records}
    for phone, record in phone_to_record.items():
        expected_for_the_phone = expected_features[phone]
        for feature_name, expected_value in expected_for_the_phone.items():
            if expected_value is not None:
                if isinstance(expected_value, list):
                    expected_value = _convert_list_feature_to_string(
                        expected_value,
                    )
                assert str(record[feature_name]) == expected_value
                continue
            assert feature_name not in record or not record[feature_name]


@pytest.mark.parametrize(
    ('with_template_file', 'file_extension'),
    [(False, ''), (True, 'xlsx'), (True, 'xlsm')],
)
@pytest.mark.project(slug='sample_project')
async def test_outgoing_calls_results_task_no_template_file(
        stq,
        stq3_context,
        stq_runner,
        mockserver,
        load_binary,
        web_app_client,
        with_template_file,
        file_extension,
        create_task,
        get_task_file,
):
    supposed_ai_questions = ['intro text', 'question']
    supposed_user_answers = ['hello!', 'answer']
    policy_titles = ['1. Intro', '2. Question', '3. End']
    errors_vs_indices = [
        (None, None),
        ('DESTINATION_NO_ANSWER', 0),
        ('ABONENT_HANGUP', 1),
    ]

    @mockserver.json_handler('/supportai-context/v1/contexts/multiple')
    async def _(_):
        contexts = [
            common.create_context_from_dialog(
                chat_id=str(idx + 1),
                ai_questions=supposed_ai_questions,
                user_answers=supposed_user_answers,
                policy_titles=policy_titles,
                error_code=error_code,
                error_code_index=error_code_index,
            )
            for idx, (error_code, error_code_index) in enumerate(
                errors_vs_indices,
            )
        ]
        return web.json_response(
            data={'contexts': contexts, 'total': len(contexts)},
        )

    if with_template_file:
        filename = f'voice_template_file.{file_extension}'
        template_file_data = load_binary(filename)

        form = aiohttp.FormData()
        form.add_field(
            name='file',
            value=template_file_data,
            filename=filename,
            content_type=FILE_EXTENSION_TO_CONTENT_TYPE[file_extension],
        )

        response_upload_file = await web_app_client.post(
            '/v1/files?project_slug=sample_project&user_id=34', data=form,
        )

        assert response_upload_file.status == 200

        response_upload_file_json = await response_upload_file.json()
        template_file_id = response_upload_file_json['file']['id']

        response_insert_config = await web_app_client.put(
            'v1/project_configs?project_slug=sample_project&user_id=34',
            json={'template_file_id': template_file_id},
        )

        assert response_insert_config.status == 200

    task = create_task(
        type_='outgoing_calls_results', params={'ref_task_id': '1'},
    )

    await calls_processing.task(
        stq3_context, 'test', task_id=task.id, task_args={'ref_task_id': '1'},
    )

    assert (
        task.status == base_task.TaskStatus.COMPLETED.value
    ), task.error_message

    assert task.file_id is not None

    file_data = get_task_file(task.file_id)

    assert file_data is not None

    workbook = openpyxl.load_workbook(io.BytesIO(file_data[1]))

    assert len(workbook.sheetnames) == (4 if with_template_file else 2)
    assert 'Raw Calls Data' in workbook
    assert 'Parsed Calls Data' in workbook

    parsed_calls_data_sheet = workbook['Parsed Calls Data']

    parsed_sheet_colnames = [
        col[0].value for col in parsed_calls_data_sheet.iter_cols()
    ]
    assert parsed_sheet_colnames[:5] == [
        'Телефон',
        'Дата',
        'Дозвонились',
        'Время звонка(сек)',
        'Время разговора (сек)',
    ]
    assert parsed_sheet_colnames[9:] == ['1. Intro', '2. Question', '3. End']

    features_names = {'one', 'two', 'three', 'four'}
    features_colnames = parsed_sheet_colnames[5:9]

    assert features_names == set(features_colnames)

    features_mask = {
        '1': (True, True, False, False),
        '2': (True, True, True, False),
        '3': (True, False, False, True),
    }

    final_scenario_indices = {'1': 2, '2': 0, '3': 1}

    chat_ids = ['1', '2', '3']
    data_rows = list(parsed_calls_data_sheet.rows)[1:]

    for data_row in parsed_calls_data_sheet.iter_rows():
        for elem in data_row:
            print(elem.value)

    for chat_id, data_row in zip(chat_ids, data_rows):
        assert data_row[0].value == 'phone' + chat_id
        if chat_id == '2':
            assert not data_row[2].value
        else:
            assert data_row[2].value
        for idx, feature_name in enumerate(features_colnames):
            if features_mask[chat_id][idx]:
                assert data_row[5 + idx].value == feature_name + '_' + chat_id
            else:
                assert data_row[5 + idx].value is None

        final_scenario_index = final_scenario_indices[chat_id]
        for idx, cell in enumerate(data_row[9:12]):
            if idx < final_scenario_index:
                assert cell.value == supposed_user_answers[idx]
            else:
                assert cell.value is None

    if with_template_file:
        assert 'Some regular data' in workbook
        assert '>>>RaNdOm DaTa<<<' in workbook
        regular_sheet = workbook['Some regular data']
        assert regular_sheet.max_row == 4
        assert regular_sheet.max_column == 3
        assert regular_sheet.cell(1, 1).value == 'first column'
        assert regular_sheet.cell(1, 3).value == 'third column'
        assert regular_sheet.cell(2, 1).value == 'Data 1.1'
        assert regular_sheet.cell(3, 2).value is None
        assert regular_sheet.cell(3, 3).value == 'data2.3'

        random_sheet = workbook['>>>RaNdOm DaTa<<<']
        assert random_sheet.max_row == 17
        assert random_sheet.max_column == 6
        assert random_sheet['C1'].value == 'C1'
        assert random_sheet.cell(7, 1).value == 777888999
        assert random_sheet['C10'].value == 'C10: nothing at C9'
        assert random_sheet['C9'].value is None
        assert random_sheet['B17'].value == 'B17: null at A5'
        assert random_sheet['A5'].value == 'null'


NEWER_THAN_TS = 10 ** 12


@pytest.mark.parametrize(
    ('direction_filter', 'newer_than_filter', 'calls_results_limit_filter'),
    [
        ('outgoing', None, None),
        ('incoming', None, None),
        (None, None, None),
        ('incoming', NEWER_THAN_TS, None),
        (None, None, 2),
    ],
)
@pytest.mark.project(slug='separate_calls_results')
async def test_separate_calls_results(
        stq3_context,
        stq_runner,
        mockserver,
        create_task,
        get_task_file,
        direction_filter,
        newer_than_filter,
        calls_results_limit_filter,
):
    project_slug = 'separate_calls_results'

    pivot_datetime = datetime.datetime.fromtimestamp(
        NEWER_THAN_TS / 1000, tz=datetime.timezone.utc,
    )
    delta = datetime.timedelta(seconds=1)
    calls_created = [
        pivot_datetime + multiple * delta for multiple in range(-3, 4, 2)
    ]

    calls = []
    for idx, (call_created, call_direction) in enumerate(
            zip(calls_created, ['incoming', 'outgoing'] * 2),
    ):
        calls.append(
            common.get_preset_call(
                project_slug,
                phone=str(idx),
                chat_id=str(idx),
                direction=db_models.CallDirection(call_direction),
                created=call_created,
                status=db_models.CallStatus.ENDED,
            ),
        )
    async with stq3_context.pg.master_pool.acquire() as conn:
        await db_models.Call.insert_bulk(stq3_context, conn, calls)

    if calls_results_limit_filter:
        expected_chat_ids = {'0', '1'}
    elif not direction_filter and not newer_than_filter:
        expected_chat_ids = {'0', '1', '2'}
    else:
        direction_to_chat_ids = {
            'incoming': {'0', '2'},
            'outgoing': {'1', '3'},
        }
        subsets = []
        if direction_filter:
            subsets.append(direction_to_chat_ids[direction_filter])
        if newer_than_filter:
            subsets.append({'2', '3'})

        expected_chat_ids = {
            chat_id
            for chat_id in {'0', '1', '2', '3'}
            if all([chat_id in subset for subset in subsets])
        }

    @mockserver.json_handler('supportai-context/v1/contexts/multiple')
    async def _(request):
        chat_ids = request.json['chat_ids']
        assert set(chat_ids) == expected_chat_ids
        contexts = [
            common.create_context_from_dialog(
                chat_id=chat_id,
                ai_questions=['hello'],
                user_answers=['oh hello'],
            )
            for chat_id in chat_ids
        ]
        return {'contexts': contexts, 'total': len(chat_ids)}

    task_args = {'results_from_separate_calls': True}
    if direction_filter:
        task_args['direction'] = direction_filter
    if newer_than_filter:
        task_args['newer_than'] = newer_than_filter
    if calls_results_limit_filter:
        task_args['calls_results_limit'] = calls_results_limit_filter

    task = create_task(type_='outgoing_calls_results', params=task_args)
    await calls_processing.task(
        stq3_context, 'test', task_id=task.id, task_args=task_args,
    )
    assert task.status == base_task.TaskStatus.COMPLETED.value
    assert task.file_id is not None
    file_data = get_task_file(task.file_id)
    assert file_data is not None

    workbook = openpyxl.load_workbook(io.BytesIO(file_data[1]))
    assert len(workbook.sheetnames) == 2
    assert 'Raw Calls Data' in workbook
    assert 'Parsed Calls Data' in workbook

    parsed_calls_data_sheet = workbook['Parsed Calls Data']

    parsed_sheet_colnames = [
        col[0].value for col in parsed_calls_data_sheet.iter_cols()
    ]
    assert parsed_sheet_colnames[:6] == [
        'Телефон',
        'Входящий/исходящий',
        'Дата',
        'Дозвонились',
        'Время звонка(сек)',
        'Время разговора (сек)',
    ]
    assert parsed_sheet_colnames[6:] == ['hello']

    parsed_sheet_first_row = [
        row[0].value for row in parsed_calls_data_sheet.iter_rows()
    ]
    assert set(parsed_sheet_first_row[1:]) == expected_chat_ids


@pytest.mark.project(id='1', slug='expose_first')
@pytest.mark.project(id='2', slug='expose_first_and_second')
async def test_features_exposure(
        stq3_context, stq_runner, mockserver, create_task, get_task_file,
):
    features_for_contexts = {
        'expose_first': {
            'chat1': [
                {'key': 'first', 'value': 'should be exposed'},
                {'key': 'second', 'value': 'should not be exposed'},
            ],
            'chat2': [{'key': 'second', 'value': 'frecking yeah'}],
        },
        'expose_first_and_second': {
            'chat1': [
                {'key': 'first', 'value': 'overwritten feature value'},
                {'key': 'second', 'value': 'one more exposed feature!'},
                {'key': 'fun_feature', 'value': 'just for fun, he-he'},
            ],
        },
    }
    expected_features_dict = {
        'expose_first': {
            'phone1': {
                'initial': 'initial1',
                'one_more_initial': None,
                'first': 'should be exposed',
                'second': None,
            },
            'phone2': {
                'initial': 'initial2',
                'one_more_initial': 'one_more_initial2',
                'first': None,
                'second': None,
            },
        },
        'expose_first_and_second': {
            'phone1': {
                'initial': 'initial',
                'first': 'overwritten feature value',
                'second': 'one more exposed feature!',
            },
        },
    }

    @mockserver.json_handler('supportai-context/v1/contexts/multiple')
    async def _(request):
        chat_ids = request.json['chat_ids']
        project_slug = request.query['project_id']
        contexts = [
            common.create_context(
                chat_id=chat_id,
                num_records=1,
                responses_features=[
                    features_for_contexts[project_slug][chat_id],
                ],
            )
            for chat_id in chat_ids
        ]
        return {'contexts': contexts, 'total': len(chat_ids)}

    task_first = create_task(project_id='1', type_='outgoing_calls_results')

    task_first_and_second = create_task(
        project_id='2', type_='outgoing_calls_results',
    )

    await stq_runner.supportai_calls_processing.call(
        task_id=4,
        args=(),
        kwargs={'task_id': task_first.id, 'task_args': {'ref_task_id': '4'}},
    )
    await stq_runner.supportai_calls_processing.call(
        task_id=5,
        args=(),
        kwargs={
            'task_id': task_first_and_second.id,
            'task_args': {'ref_task_id': '5'},
        },
    )

    assert task_first.status == base_task.TaskStatus.COMPLETED.value
    assert task_first_and_second.status == base_task.TaskStatus.COMPLETED.value

    file_first_data = get_task_file(task_first.file_id)
    file_first_and_second_data = get_task_file(task_first_and_second.file_id)

    builtin_columns = list(
        outgoing_calls_results_task.PARAM_NAME_TO_COLNAME.values(),
    )
    builtin_columns.pop(1)
    columns_first = builtin_columns + ['first']
    columns_first_and_second = builtin_columns + ['first', 'second']

    _check_table_header_beginning(file_first_data[1], columns_first)
    try:
        _check_table_header_beginning(
            file_first_data[1], columns_first_and_second,
        )
        raise 'hello!'
    except AssertionError:
        pass

    _check_table_header_beginning(
        file_first_and_second_data[1], columns_first_and_second,
    )

    _check_features_in_file(
        file_first_data[1], expected_features_dict['expose_first'],
    )
    _check_features_in_file(
        file_first_and_second_data[1],
        expected_features_dict['expose_first_and_second'],
    )


@pytest.mark.project(slug='expose_first_second_and_list')
async def test_exposure_features_through_history(
        stq3_context, stq_runner, mockserver, create_task, get_task_file,
):
    first_feature_layers = [None, 'initial', None, None, 'changed']
    second_feature_layers = [None, None, 'stable', None, None]
    list_str_feature_layers = [None, ['one'], ['one', 'two'], None, None]
    list_float_feature_layers = [[0.0], [], None, None, [1.6]]
    not_related_feature_layers = [
        'initial',
        None,
        None,
        'changed',
        'yet_again_changed',
    ]
    first_tuples = [('first', first) for first in first_feature_layers]
    second_tuples = [('second', second) for second in second_feature_layers]
    list_str_tuples = [
        ('list_str', list_str) for list_str in list_str_feature_layers
    ]
    list_float_tuples = [
        ('list_float', list_float) for list_float in list_float_feature_layers
    ]
    not_related_tuples = [
        ('not_related', not_related)
        for not_related in not_related_feature_layers
    ]

    (*generations,) = zip(
        first_tuples,
        second_tuples,
        list_str_tuples,
        list_float_tuples,
        not_related_tuples,
    )

    responses_features = [
        [
            {'key': key, 'value': value}
            for key, value in generation
            if value is not None
        ]
        for generation in generations
    ]

    expected_features = {
        'phone1': {
            'first': 'changed',
            'second': 'stable',
            'list_str': ['one', 'two'],
            'list_float': [1.6],
        },
    }

    task = create_task(type_='outgoing_calls_results')

    @mockserver.json_handler('supportai-context/v1/contexts/multiple')
    async def _(request):
        chat_id = request.json['chat_ids'][0]
        contexts = [
            common.create_context(
                chat_id, responses_features=responses_features,
            ),
        ]
        return {'contexts': contexts, 'total': 1}

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id,
        args=(),
        kwargs={'task_id': task.id, 'task_args': {'ref_task_id': '8'}},
    )

    file_data = get_task_file(task.file_id)

    assert task.status == base_task.TaskStatus.COMPLETED.value

    _check_features_in_file(file_data[1], expected_features)


async def test_collect_calls_data_with_voice_synthesis(
        stq3_context, stq_runner, mockserver, create_task, get_task_file,
):
    @mockserver.json_handler('supportai-context/v1/contexts/multiple')
    async def _(request):
        chat_id = request.json['chat_ids'][0]
        return {
            'contexts': [
                common.create_context_from_dialog(
                    chat_id=chat_id,
                    ai_questions=['intro phrase', 'question'],
                    user_answers=['hello!', 'answer'],
                    policy_titles=['1. Intro', '2. Question', '3. End'],
                    tags=['speak'],
                ),
            ],
            'total': 1,
        }

    task = create_task(type_='outgoing_calls_results')

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id,
        args=(),
        kwargs={'task_id': task.id, 'task_args': {'ref_task_id': '6'}},
    )

    assert task.status == base_task.TaskStatus.COMPLETED.value

    file_data = get_task_file(task.file_id)

    result_wb = openpyxl.load_workbook(io.BytesIO(file_data[1]))

    raw_sheet = result_wb['Raw Calls Data']
    assert raw_sheet.cell(2, 3).value == 'intro phrase'
    assert raw_sheet.cell(2, 4).value == 'hello!'
    assert raw_sheet.cell(2, 5).value == 'question'
    assert raw_sheet.cell(2, 6).value == 'answer'
    assert raw_sheet.cell(2, 7).value is None

    parsed_sheet = result_wb['Parsed Calls Data']
    assert parsed_sheet.cell(1, 6).value == '1. Intro'
    assert parsed_sheet.cell(2, 6).value == 'hello!'
    assert parsed_sheet.cell(1, 7).value == '2. Question'
    assert parsed_sheet.cell(2, 7).value == 'answer'
    assert parsed_sheet.cell(1, 8).value == '3. End'
    assert parsed_sheet.cell(2, 8).value is None


async def test_order_of_policy_slugs(
        stq3_context, stq_runner, mockserver, create_task, get_task_file,
):
    expected_policies_order = [
        '1. With dot',
        '1.1.1 No dot',
        '1.1.21. Adding dot does not matter',
        '1.11.1 Comparing numbers separately',
        '1.21. Comparing vectors lexicographically',
        '1.100. Comparing numbers numerically',
        '2. Just for fun!',
        '2.1 Should be before the next title',
        '2 1 3 You may use whitespaces as well to separate numbers',
        '2-=-++  2 You may use any symbol to separate numbers',
        '25 one number and no dots',
        '200 order is still numeric',
    ]
    policy_titles = copy.copy(expected_policies_order)
    random.shuffle(policy_titles)

    @mockserver.json_handler('supportai-context/v1/contexts/multiple')
    # pylint: disable=unused-variable
    async def handler(request):
        records = [
            common.create_context_record(policy_title=policy_title)
            for policy_title in policy_titles
        ]
        return {
            'contexts': [
                {
                    'created_at': str(datetime.datetime.now().astimezone()),
                    'chat_id': request.json['chat_ids'][0],
                    'records': records,
                },
            ],
            'total': 1,
        }

    task = create_task(type_='outgoing_calls_results')

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id,
        args=(),
        kwargs={'task_id': task.id, 'task_args': {'ref_task_id': '6'}},
    )

    file_data = get_task_file(task.file_id)

    result_wb = openpyxl.load_workbook(io.BytesIO(file_data[1]))
    parsed_sheet = result_wb['Parsed Calls Data']
    for idx, expected_policy_title in enumerate(expected_policies_order):
        assert parsed_sheet.cell(1, 6 + idx).value == expected_policy_title


@pytest.mark.parametrize(
    ('version', 'client_error'), [('v1', False), ('v2', False), ('v2', True)],
)
async def test_results_with_version(
        stq3_context,
        stq_runner,
        mockserver,
        version,
        client_error,
        create_task,
        get_task_file,
):
    scenario_titles_answered = [
        '1. Some question',
        '2. Another question',
        '3. End',
    ]
    user_answers = ['Some answer', 'Some another answer']
    title_to_answer = {
        title: answer
        for title, answer in zip(scenario_titles_answered, user_answers)
    }

    scenario_titles_v1 = scenario_titles_answered + [
        f'10.{i} Title' for i in range(5)
    ]
    new_scenario_titles = [f'20.{i} Title' for i in range(5)]
    scenario_titles_v2 = scenario_titles_v1 + new_scenario_titles
    version_to_scenarios = {'v1': scenario_titles_v1, 'v2': scenario_titles_v2}
    scenario_titles_unavailable = [
        f'0.{i} WOULD BREAK THE SYSTEM' for i in range(2)
    ]

    @mockserver.json_handler('supportai-context/v1/contexts/multiple')
    async def _(request):
        chat_id = request.json['chat_ids'][0]
        return {
            'contexts': [
                common.create_context_from_dialog(
                    chat_id=chat_id,
                    ai_questions=['', ''],
                    user_answers=user_answers,
                    version=version,
                    policy_titles=scenario_titles_answered,
                ),
            ],
            'total': 1,
        }

    @mockserver.json_handler('supportai/supportai/v1/scenarios')
    async def _(request):
        if client_error:
            return errors.HTTPBadRequest()

        if int(request.query['offset']) > 0:
            return []

        scenario_titles_available = version_to_scenarios.get(version)

        def get_scenario(scenario_title, available):
            return {
                'id': '',
                'title': scenario_title,
                'rule': {'parts': ['']},
                'available': available,
                'actions': [{'type': 'response'}],
            }

        scenarios = (
            [
                get_scenario(scenario_title, available=True)
                for scenario_title in scenario_titles_available
            ]
            + [
                get_scenario(scenario_title, available=False)
                for scenario_title in scenario_titles_unavailable
            ]
        )

        return {'scenarios': scenarios}

    task = create_task(type_='outgoing_calls_results')

    await stq_runner.supportai_calls_processing.call(
        task_id=task.id,
        args=(),
        kwargs={'task_id': task.id, 'task_args': {'ref_task_id': '6'}},
    )

    assert (
        task.status == base_task.TaskStatus.COMPLETED.value
    ), f'Incorrect status, error_message: {task.error_message}'

    assert task.file_id is not None

    file_data = get_task_file(task.file_id)

    assert file_data is not None

    workbook = openpyxl.load_workbook(io.BytesIO(file_data[1]))
    parsed_data_sheet = workbook['Parsed Calls Data']

    if client_error:
        expected_scenarios = scenario_titles_answered
    else:
        expected_scenarios = version_to_scenarios[version]

    for idx, scenario_title in enumerate(expected_scenarios):
        expected_answer = title_to_answer.get(scenario_title, '')
        assert parsed_data_sheet.cell(1, idx + 6).value == scenario_title
        if expected_answer:
            assert parsed_data_sheet.cell(2, idx + 6).value == expected_answer
        else:
            assert not parsed_data_sheet.cell(2, idx + 6).value

    assert not parsed_data_sheet.cell(1, len(expected_scenarios) + 6).value
