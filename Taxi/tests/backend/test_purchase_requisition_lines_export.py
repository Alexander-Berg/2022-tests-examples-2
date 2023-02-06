import xlrd
import datetime
import openpyxl
import base64
from io import BytesIO

from freezegun import freeze_time
from odoo import fields
from odoo.tests import tagged,  HttpSavepointCase
from odoo.addons.lavka.backend.wizzard.import_requisition_lines_wizzard import example_line


@tagged('lavka', 'autoorder', 'purchase_requisition_lines', 'autoorder_export')
class TestPurchaseLinesDownload(HttpSavepointCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()

        cls.test_user = cls.env['res.users'].create({
            'login': 'test',
            'password': 'test',
            'partner_id': cls.env['res.partner'].create({
                'name': 'Test'
            }).id,
            'groups_id': [cls.env.ref('lavka.group_catman').id],
        })

        cls.wh_tag1 = cls.env['stock.warehouse.tag'].create(
            {
                'type': 'geo',
                'name': 'Paris 1',
            }
        )

        cls.p1 = cls.env['product.product'].create(
            {
                'name': 'p-1',
                'default_code': '1',
                'wms_id': '1',
            }
        )

        cls.p2 = cls.env['product.product'].create(
            {
                'name': 'p-2',
                'default_code': '2',
                'wms_id': '2',
            }
        )

        cls.p3 = cls.env['product.product'].create(
            {
                'name': 'p-3',
                'default_code': '3',
                'wms_id': '3',
            }
        )

        cls.p4 = cls.env['product.product'].create(
            {
                'name': 'p-4',
                'default_code': '4',
                'wms_id': '4',
            }
        )

        cls.v1 = cls.env['res.partner'].create(
            {
                'name': 'v-1',
                'is_company': True,
                'supplier_rank': 1,
            }
        )

        cls.v2 = cls.env['res.partner'].create(
            {
                'name': 'v-1',
                'is_company': True,
                'supplier_rank': 1,
            }
        )

        cls.pr1 = cls.env['purchase.requisition'].create(
            {
                'vendor_id': cls.v1.id,
                'state': 'ongoing',
                'warehouse_tag_ids': [cls.wh_tag1.id],
            }
        )

        cls.pr2 = cls.env['purchase.requisition'].create(
            {
                'vendor_id': cls.v2.id,
                'state': 'ongoing',
                'warehouse_tag_ids': [cls.wh_tag1.id],
            }
        )

        products_lines_data = (
            (cls.pr1, cls.p1, 100, 1, 1, True, True, True, True,),
            (cls.pr1, cls.p2, 200, 2, 2, True, True, True, True,),
            (cls.pr1, cls.p3, 200, 2, 2, True, False, False, True,),
            (cls.pr1, cls.p4, 200, 2, 2, True, False, False, False,),
        )

        cls.env['purchase.requisition.line'].create(
            [
                {
                    'requisition_id': pr.id,
                    'product_id': p.id,
                    'start_date': fields.Datetime.today(),
                    'tax_id': p.supplier_taxes_id.id,
                    'product_uom_id': p.uom_id.id,
                    'price_unit': price,
                    'product_qty': qty,
                    'product_code': f'code-{p.default_code}',
                    'product_name': f'name-{p.default_code}',
                    'qty_multiple': qty_multiple,
                    'approve_price': approve_price,
                    'approve_tax': approve_tax,
                    'approve': approve,
                    'active': active,
                }
                for (pr, p, price, qty, qty_multiple,
                     approve_price, approve_tax, approve, active)
                in products_lines_data
            ]
        )

        products_lines_data_2 = (
            (cls.pr2, cls.p1, 200, 2, 2, True, False, False, True,),
            (cls.pr2, cls.p2, 200, 2, 2, True, False, False, False,),
        )

        cls.env['purchase.requisition.line'].create(
            [
                {
                    'requisition_id': pr.id,
                    'product_id': p.id,
                    'start_date': fields.Datetime.today(),
                    'end_date': fields.Datetime.today() + datetime.timedelta(days=1),
                    'tax_id': p.supplier_taxes_id.id,
                    'product_uom_id': p.uom_id.id,
                    'price_unit': price,
                    'product_qty': qty,
                    'product_code': f'code-{p.default_code}',
                    'product_name': f'name-{p.default_code}',
                    'qty_multiple': qty_multiple,
                    'approve_price': approve_price,
                    'approve_tax': approve_tax,
                    'approve': approve,
                    'active': active,
                }
                for (pr, p, price, qty, qty_multiple,
                     approve_price, approve_tax, approve, active)
                in products_lines_data_2
            ]
        )

        products_lines_data_3 = (
            (cls.pr2, cls.p3, 200, 2, 2, True, False, False, False,),
        )
        cls.env['purchase.requisition.line'].create(
            [
                {
                    'requisition_id': pr.id,
                    'product_id': p.id,
                    'start_date': fields.Datetime.today(),
                    'end_date': fields.Datetime.today() + datetime.timedelta(days=1),
                    'tax_id': p.supplier_taxes_id.id,
                    'product_uom_id': p.uom_id.id,
                    'price_unit': price,
                    'product_qty': qty,
                    'product_code': f'code-{p.default_code}',
                    'product_name': f'name-{p.default_code}',
                    'qty_multiple': qty_multiple,
                    'approve_price': approve_price,
                    'approve_tax': approve_tax,
                    'approve': approve,
                    'active': active,
                    'create_date': fields.Datetime.today() + datetime.timedelta(days=1),
                }
                for (pr, p, price, qty, qty_multiple,
                     approve_price, approve_tax, approve, active)
                in products_lines_data_3
            ]
        )

    def _headers(self):
        return [
            'default_code',
            'description',
            'start_date',
            'end_date',
            'qty_multiple',
            'qty_in_box',
            'unit_price',
            'tax_value',
            'vendor_product_code',
            'vendor_product_name',
        ]

    def test_export_lines(self):
        today_dt = datetime.datetime(*datetime.date.today().timetuple()[:5])
        test_rows = [
            self._headers(),
            [
                '1',
                None,
                today_dt,
                None,
                1.0,
                1.0,
                100.0,
                15.0,
                'code-1',
                'name-1',
            ],
            [
                '2',
                None,
                today_dt,
                None,
                2.0,
                1.0,
                200.0,
                15.0,
                'code-2',
                'name-2',
            ],  # 3,4 цены недействующие, поэтому не учитываются в выгрузке
        ]

        session = self.authenticate('test', 'test')
        with freeze_time(today_dt):
            response = self.url_open(
                '/web/download/purchase-lines?req_id=%s' % self.pr1.id,
                headers={'session_id': session['session_token']},
            )

        self.assertEqual(response.status_code, 200)

        xlsx_file = BytesIO(response.content)
        wb_obj = openpyxl.load_workbook(xlsx_file, data_only=False)
        sheet = wb_obj.active

        rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
        self.assertEqual(rows, test_rows)


    def test_export_lines_old_prices(self):
        today_dt = datetime.datetime(*datetime.date.today().timetuple()[:5])
        test_rows = [
            self._headers(),
            [
                '3',
                None,
                today_dt,
                None,
                2.0,
                1.0,
                200,
                15,
                'code-3',
                'name-3',
            ], #  последняя загруженная цена
        ]

        session = self.authenticate('test', 'test')
        with freeze_time(today_dt + datetime.timedelta(days=2)):
            response = self.url_open(
                '/web/download/purchase-lines?req_id=%s' % self.pr2.id,
                headers={'session_id': session['session_token']},
            )

        self.assertEqual(response.status_code, 200)

        xlsx_file = BytesIO(response.content)
        wb_obj = openpyxl.load_workbook(xlsx_file, data_only=False)
        sheet = wb_obj.active

        rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
        self.assertEqual(rows, test_rows)

    def test_export_lines_with_bad_req_id(self):
        test_rows = [
            self._headers(),
            list(example_line.values()),
        ]

        session = self.authenticate('test', 'test')
        response = self.url_open(
            '/web/download/purchase-lines?req_id=%s' % -1,
            headers={'session_id': session['session_token']},
        )

        self.assertEqual(response.status_code, 200)

        xlsx_file = BytesIO(response.content)
        wb_obj = openpyxl.load_workbook(xlsx_file, data_only=True)
        sheet = wb_obj.active

        rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
        self.assertEqual(rows, test_rows)
