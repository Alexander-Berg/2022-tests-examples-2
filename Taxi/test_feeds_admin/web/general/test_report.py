import io

import openpyxl
import pytest

from test_feeds_admin import const


@pytest.mark.pgsql('feeds_admin', files=['report.sql'])
@pytest.mark.config(
    FEEDS_ADMIN_REPORTS={
        'contractor-marketplace': {
            'test': [
                {'path': 'feed.feed_id', 'name': 'feed_id'},
                {'path': 'feed.created', 'name': 'Дата создания'},
                {'path': 'payload.partner.name', 'name': 'partner_name'},
                {'path': 'unknown', 'name': 'Пустой столбец'},
                {'path': 'schedule.first_start_at', 'name': 'Начало'},
                {'path': 'schedule.next_finish_at', 'name': 'Конец'},
                {'path': 'recipients.type.tag', 'name': 'Теги'},
            ],
        },
    },
)
async def test_contractor_marketplace(
        web_app_client, patch, mock_feeds,
):  # pylint: disable=W0612
    response = await web_app_client.get(
        '/v1/report',
        params={'service': 'contractor-marketplace', 'report': 'test'},
    )
    assert response.status == 200
    content = await response.read()
    workbook = openpyxl.load_workbook(filename=io.BytesIO(content))
    sheet = workbook.active
    assert sheet['A1'].value == 'feed_id'
    assert sheet['A2'].value == const.UUID_1
    has_tags = False
    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == 'Теги':
            has_tags = True
            assert sheet.cell(row=2, column=i).value == 'tag1, tag2'
    assert has_tags
